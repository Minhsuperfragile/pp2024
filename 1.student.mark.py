from openpyxl import load_workbook
from random import randint

studentDictionary = {
    "students": [],
    "numberOfStudents": 0,
    "numberOfCourses": 0,
    "courses": []
}

#region Input Functions
def getNumberOfStudent(n:int):
    studentDictionary["numberOfStudents"] = n

def getStudentInformation(name:str, Id:str,Dob:str):
    if len(studentDictionary["students"]) == studentDictionary["numberOfStudents"]:
        studentDictionary["numberOfStudents"] += 1
    student = {
        "name":name,
        "id":Id,
        "dob":Dob
    }
    studentDictionary["students"].append(student)

def getNumberOfCourses(n:int):
    studentDictionary["numberOfCourses"] = n

def getCourseInfomation(Id:str,name:str):
    if len(studentDictionary["courses"]) == studentDictionary["numberOfCourses"]:
        studentDictionary["numberOfCourses"] += 1
    course = {
        "name":name,
        "id":Id,
        "mark": []
    }
    for i in range(0,studentDictionary["numberOfStudents"]):
        course['mark'].append(0)
    studentDictionary["courses"].append(course)

def getStudentMark(studentID:str,courseID:str,mark:float):
    index = 0
    for student in studentDictionary["students"]:
        if student["id"] == studentID:
            break
        index += 1

    for course in studentDictionary["courses"]:
        if course["id"] == courseID:
            course["mark"][index] = mark
#endregion

#region Listing Functions
def listCourses() -> None:
    for course in studentDictionary["courses"]:
        print(f"Course Name: {course['name']} - ID: {course['id']}")

def listStudent() -> None:
    for student in studentDictionary["students"]:
        print(f"{student['name']} - {student['id']} - {student['dob']}")

def listMark(courseID:str) -> None:
    for course in studentDictionary["courses"]:
        if courseID == course["id"]:
            for i in range(0,studentDictionary["numberOfStudents"]):
                print(f"{studentDictionary['students'][i]['name']} got {course['mark'][i]:.2f}")
#endregion

def main() -> None:  
    path = __file__[:-17]

    wb = load_workbook(path + "studentData.xlsx")
    ws = wb.active

    getNumberOfCourses(3)
    getNumberOfStudent(469)

    #input courses
    getCourseInfomation("APP","Advanced Programming with Python")
    getCourseInfomation("OOP","Object Oriented Programming")
    getCourseInfomation("ADS","Algorithm and Data Structure")

    #input students information and mark
    for i in range(2,471):
        name = ws.cell(row=i,column=2).value +" "+ ws.cell(row=i,column=3).value
        
        dob = str(ws.cell(row=i,column=4).value)

        if len(dob) > 10:
            dob = dob[:-9]

        id = ws.cell(row=i,column=1).value
        getStudentInformation(name, id, dob)
        getStudentMark(id,"OOP",randint(5,18))
        getStudentMark(id,"APP",randint(5,18))
        getStudentMark(id,"ADS",randint(5,18))

    print("Students:\n")
    listStudent()
    print("Courses:\n")
    listCourses()
    print("Marks:\n")
    listMark("OOP")
    print("--------------------\n")
    listMark("ADS")
    print("--------------------:\n")
    listMark("APP")

main()