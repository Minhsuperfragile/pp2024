from openpyxl import load_workbook
from math import floor
from random import random, seed
from domain import CourseClass,StudentClass
from os import name as osn 
import pandas as pd

def getPath() -> str:
    if osn == 'nt':
        path = __file__[:-len( __file__.split('\\')[-1])]
    else:
        path = __file__[:-len( __file__.split('/')[-1])]
    return path

def decompress(studentList:list,courseList:list) -> int:
    try:
        students = pd.read_csv(getPath() + 'students.dat')
        
        #courses
        numberOfCourse = int(students.head().iloc[0,5])
        cols = list(students.columns)
        
        for i in range(numberOfCourse):
            IdAndCredit = students.head().iloc[i,4].split('@')
            courseList.append(CourseClass.Course(
                cols[i-numberOfCourse],
                IdAndCredit[0],
                int(IdAndCredit[1])
                ))

        #students
        for i in range(len(students)):
            # print(courseList)
            marks = []
            for k in range(7,7+numberOfCourse):
                marks.append(int(students.iat[i,k]))

            studentList.append(
                StudentClass.Student(
                    students.iat[i,1],
                    students.iat[i,2],
                    students.iat[i,3],
                    courseList, marks
                )
            )

        return 0
    except:
        print('failed to open students.dat')
        return 1

def sampleCourses() -> list:
    app = CourseClass.Course("Advanced Programming with Python","APP",4)
    ads = CourseClass.Course("Algorithm and Data Structure","ADS",3)
    oop = CourseClass.Course("Object Oriented Programming","OOP",3)
    return [app,ads,oop]

def sampleStudents(courseList:list) -> list:
    path = getPath()

    initSeed = 1573
    wb = load_workbook(path + "studentData.xlsx")
    ws = wb.active

    studentList = []
    for i in range(2,471):
        name = ws.cell(row=i,column=2).value +" "+ ws.cell(row=i,column=3).value
        
        dob = str(ws.cell(row=i,column=4).value)

        if len(dob) > 10:
            dob = dob[:-9]

        id = ws.cell(row=i,column=1).value

        seed(initSeed)
        initSeed+= floor(random()*1234)
        markList = [random()*20, random()*20, random()*20]
        
        studentList += [StudentClass.Student(name,id,dob,courseList,markList)]
        
    return studentList

if __name__ == '__main__':
    stList = []
    csList = []
    decompress(stList,csList)

