from openpyxl import load_workbook
from math import floor
from random import random, seed, shuffle
from os import system as st, name as osn
import tkinter as tk
from tkinter import ttk

#region classes and functions
class Course():
    __courseName:str
    __courseID:str
    __credit:int

    def __init__(self, courseName:str, courseID:str, courseCredit:int ) -> None:
        self.__courseName = courseName
        self.__courseID = courseID
        self.__credit = courseCredit

    def get__id(self) -> str:
        return self.__courseID
    
    def get__name(self) -> str:
        return self.__courseName 
    
    def get__credit(self) -> int:
        return self.__credit

    def printOut(self) -> None:
        print(f'Course: {self.__courseName}, ID: {self.__courseID}\n')

class Student():
    __name:str
    __id:str
    __dob:str
    __course:list|Course
    __mark:list|float
    __gpa:float
    __numberOfCourse:int

    def addCourse(self, course:Course, mark:float = 0) -> None:
        self.__course.append(course)
        self.__mark.append(mark)
        self.__numberOfCourse += 1

    def __calculateGPA(self) -> float:
        gpa:float = 0
        totalCredit:int = 0
        for i in range(0,self.__numberOfCourse):
            course = self.__course[i]
            mark = self.__mark[i]
            totalCredit += course.get__credit()
            gpa += course.get__credit() * mark 
        return gpa/totalCredit

    def __init__(self,name:str,id:str,dob:str,course:Course|list,mark:float|list=0) -> None:
        if type(course) == list:
            noc = len(course)
            self.__numberOfCourse = noc
            if type(mark) != list:
                mark = [mark]
                for i in range(1,noc):
                    mark.append(0)
            elif len(mark) < noc:
                for i in range(len(mark),noc):
                    mark.append(0)
        else:
            self.__numberOfCourse = 0
            if type(mark) == list:
                self.addCourse(course,mark[0])
            else:
                self.addCourse(course,mark)

        for i in range(0,len(mark)):
            mark[i] = floor(mark[i])
        
        self.__course = course
        self.__mark = mark
        self.__gpa = self.__calculateGPA()
        self.__name = name
        self.__id = id
        self.__dob = dob

    def set__mark(self, mark:list) -> None:
        self.__mark = mark

    def get__name(self) -> str:
        return self.__name

    def get__dob(self) -> str:
        return self.__dob
    
    def get__id(self) -> str:
        return self.__id
    
    def get__gpa(self) -> float:
        return self.__gpa
    
    def get__mark(self) -> list:
        return self.__mark

    def printOut(self) -> None:
        print(f'{self.__name}, {self.__id}, {self.__dob} has GPA: {self.__gpa} and involved these courses:')
        for i in range(0,self.__numberOfCourse):
            print(f'{self.__course[i].get__name()}: {self.__mark[i]}')

def sampleCourses() -> list:
    app = Course("Advanced Programming with Python","APP",4)
    ads = Course("Algorithm and Data Structure","ADS",3)
    oop = Course("Object Oriented Programming","OOP",3)
    return [app,ads,oop]

def sampleStudents(courseList:list) -> list:
    path = __file__[:-len(__file__.split('\\')[-1])]
    initSeed = 1573
    wb = load_workbook(path + "studentData.xlsx")
    ws = wb.active

    studentList = []
    for i in range(2,471):
        name = ws.cell(row=i,column=2).value +" "+ ws.cell(row=i,column=3).value
        
        dob = str(ws.cell(row=i,column=4).value)

        if len(dob) > 10:
            dob = dob[:-9]

        id = ws.cell(row=i,column=1).value

        seed(initSeed)
        initSeed+= floor(random()*1234)
        markList = [random()*20, random()*20, random()*20]
        
        studentList += [Student(name,id,dob,courseList,markList)]
        
    return studentList

def sortGPA(studentList:list) -> None:
    n = len(studentList)

    for i in range(n): #bubble sort
        swapped = False
        for j in range(0,n-i-1):
            if studentList[j].get__gpa() < studentList[j+1].get__gpa():
                studentList[j], studentList[j+1] = studentList[j+1], studentList[j]
                swapped = True
        if swapped == False:
            break

def listAllOut(students:list)->None:
    for student in students:
        student.printOut()

def clear() -> None:
    st('cls' if osn == 'nt' else 'clear')
#endregion
    
#region button function
studentList = sampleStudents(sampleCourses())

def sortButtonCmd() -> None:
    sortGPA(studentList)

def listAllOutCmd() -> None:
    listAllOut(studentList)

def shuffleButtonCmd() -> None:
    shuffle(studentList)
#endregion
    
#region window UI
window = tk.Tk()
window.title('Cursed UI')
window.geometry('300x200')
window.resizable(False, False)

#ttk style
style = ttk.Style()
style.theme_use('alt')
style.configure('TButton', background = '#FFBE98', foreground = 'black', width = 20, borderwidth=1, focusthickness=3, focuscolor='none')
style.map('TButton', background=[('active','#FFD1DC')])

#ttk label
label = ttk.Label(master= window,text= 'student management terminal controller')
label.pack()

#ttk button
refreshButton = ttk.Button(master= window, text= 'refresh',command= clear)
refreshButton.pack(ipadx=7,ipady=7)

sortButton = ttk.Button(master= window, text= 'sort',command= sortButtonCmd)
sortButton.pack(ipadx=7,ipady=7)

shuffleButton = ttk.Button(master= window, text= 'shuffle',command= shuffleButtonCmd)
shuffleButton.pack(ipadx=7,ipady=7)

printButton = ttk.Button(master= window, text= 'list',command= listAllOutCmd)
printButton.pack(ipadx=7,ipady=7)

window.mainloop()
#endregion