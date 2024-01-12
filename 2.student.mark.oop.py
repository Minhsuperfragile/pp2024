from openpyxl import load_workbook
from random import randint

class Course():
    __courseName:str
    __courseID:str

    def __init__(self, courseName:str, courseID:str ) -> None:
        self.__courseName = courseName
        self.__courseID = courseID

    def get__id(self) -> str:
        return self.__courseID
    
    def get__name(self) -> str:
        return self.__courseName 
    
    def printOut(self) -> None:
        print(f'Course: {self.__courseName}, ID: {self.__courseID}\n')

class Student():
    __name:str
    __id:str
    __dob:str
    __courses:dict = {
        "course": [],
        "mark": []
    }

    def addCourse(self, course:Course, mark:float = 0):
        self.__courses["course"].append(course)
        self.__courses["mark"].append(mark)

    def __init__(self,name:str,id:str,dob:str,course:Course|list,mark:float|list=0) -> None:
        if type(course) == list:
            noc = len(course)

            if type(mark) != list:
                mark = [mark]
                for i in range(1,noc):
                    mark.append(0)
            elif len(mark) < noc:
                for i in range(len(mark),noc):
                    mark.append(0)

        self.__courses["course"] = course
        self.__courses["mark"] = mark
        self.__name = name
        self.__id = id
        self.__dob = dob

    def set__courses(self, c:dict) -> None:
        self.__courses = c

    def get__name(self) -> str:
        return self.__name

    def get__dob(self) -> str:
        return self.__dob
    
    def get__id(self) -> str:
        return self.__id

    def printOut(self) -> None:
        print(f'{self.__name}, {self.__id}, {self.__dob} involved these courses:')
        for i in range(0,len(self.__courses["course"])):
            print(f'{self.__courses["course"][i].get__name()}: {self.__courses["mark"][i]}')

def main():
    path = __file__[:-21]

    wb = load_workbook(path + "studentData.xlsx")
    ws = wb.active

    app = Course("Advanced Programming with Python","APP")
    ads = Course("Algorithm and Data Structure","ADS")
    oop = Course("Object Oriented Programming","OOP")
    courseList = [app,ads,oop]
    studentList = []

    for i in range(2,471):
        name = ws.cell(row=i,column=2).value +" "+ ws.cell(row=i,column=3).value
        
        dob = str(ws.cell(row=i,column=4).value)

        if len(dob) > 10:
            dob = dob[:-9]

        id = ws.cell(row=i,column=1).value

        markList = [0,0,0]
        markList[0] = randint(5,18)
        markList[1] = randint(5,18)
        markList[2] = randint(5,18)

        student = Student(name,id,dob,courseList,markList)
        studentList.append(student)
        student.printOut()
        
main()   